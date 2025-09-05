import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import glob
from pathlib import Path
from datetime import datetime

class MultipleJSONToSingleExcelConverter:
    def __init__(self):
        self.json_data = {}  # Dictionary to store JSON data from multiple files
        self.combined_df = None  # Single DataFrame for all data
    
    def load_multiple_json_files(self, file_paths):
        """Load multiple JSON files"""
        successful_loads = 0
        
        for file_path in file_paths:
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    filename = Path(file_path).stem
                    self.json_data[filename] = data
                    print(f"✅ Successfully loaded: {file_path}")
                    successful_loads += 1
            except FileNotFoundError:
                print(f"❌ File not found: {file_path}")
            except json.JSONDecodeError as e:
                print(f"❌ Invalid JSON in {file_path}: {e}")
            except Exception as e:
                print(f"❌ Error loading {file_path}: {e}")
        
        print(f"📊 Total files loaded: {successful_loads}")
        return successful_loads > 0
    
    def load_from_directory(self, directory_path, pattern="*.json"):
        """Load all JSON files from a directory"""
        json_files = glob.glob(os.path.join(directory_path, pattern))
        if not json_files:
            print(f"❌ No JSON files found in {directory_path}")
            return False
        
        print(f"Found {len(json_files)} JSON files in {directory_path}")
        return self.load_multiple_json_files(json_files)
    
    def combine_json_data(self, add_source_column=True):
        """Combine all JSON data into a single DataFrame"""
        if not self.json_data:
            print("❌ No JSON data loaded")
            return False
        
        all_dataframes = []
        
        for filename, data in self.json_data.items():
            try:
                # Convert JSON to DataFrame
                if isinstance(data, list):
                    # JSON is an array of objects
                    df = pd.json_normalize(data)
                elif isinstance(data, dict):
                    # Handle different dict structures
                    if any(isinstance(v, list) for v in data.values()):
                        # Find the first list and use it
                        for key, value in data.items():
                            if isinstance(value, list):
                                df = pd.json_normalize(value)
                                break
                    else:
                        # Single object - convert to single row
                        df = pd.json_normalize([data])
                else:
                    print(f"⚠️ Skipping {filename}: Unsupported data structure")
                    continue
                
                # Add source column to track which file the data came from
                if add_source_column:
                    df['source_file'] = filename
                
                all_dataframes.append(df)
                print(f"✅ Processed {filename}: {len(df)} rows, {len(df.columns)} columns")
                
            except Exception as e:
                print(f"❌ Error processing {filename}: {e}")
                continue
        
        if not all_dataframes:
            print("❌ No data could be processed")
            return False
        
        try:
            # Combine all DataFrames
            self.combined_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
            
            # Fill NaN values with empty strings for better Excel output
            self.combined_df = self.combined_df.fillna('')
            
            print(f"✅ Combined data: {len(self.combined_df)} total rows, {len(self.combined_df.columns)} columns")
            return True
            
        except Exception as e:
            print(f"❌ Error combining DataFrames: {e}")
            return False
    
    def save_to_excel_simple(self, output_file):
        """Save combined DataFrame to Excel (simple method)"""
        if self.combined_df is None:
            print("❌ No combined data available")
            return False
        
        try:
            self.combined_df.to_excel(output_file, index=False, engine='openpyxl')
            print(f"✅ Successfully saved to {output_file}")
            return True
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")
            return False
    
    def save_to_excel_styled(self, output_file):
        """Save combined DataFrame to Excel with styling"""
        if self.combined_df is None:
            print("❌ No combined data available")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Combined JSON Data"
            
            # Add data to worksheet
            for r in dataframe_to_rows(self.combined_df, index=False, header=True):
                ws.append(r)
            
            # Style the header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add alternating row colors
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            for row_num in range(2, ws.max_row + 1):
                if row_num % 2 == 0:
                    for cell in ws[row_num]:
                        cell.fill = light_fill
            
            # Highlight source_file column if it exists
            if 'source_file' in self.combined_df.columns:
                source_col_idx = list(self.combined_df.columns).index('source_file') + 1
                source_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=source_col_idx).fill = source_fill
            
            wb.save(output_file)
            print(f"✅ Successfully saved styled Excel file to {output_file}")
            return True
        except Exception as e:
            print(f"❌ Error saving styled Excel: {e}")
            return False
    
    def preview_combined_data(self, rows=10):
        """Preview the combined DataFrame"""
        if self.combined_df is None:
            print("❌ No combined data available")
            return
        
        print(f"\n📊 Combined Data Preview (first {rows} rows):")
        print("=" * 80)
        print(self.combined_df.head(rows).to_string())
        print(f"\nTotal rows: {len(self.combined_df)}")
        print(f"Total columns: {len(self.combined_df.columns)}")
        print(f"Columns: {list(self.combined_df.columns)}")
        
        # Show data distribution by source file if available
        if 'source_file' in self.combined_df.columns:
            print(f"\nData distribution by source file:")
            print(self.combined_df['source_file'].value_counts().to_string())
    
    def get_summary_stats(self):
        """Get summary statistics of the combined data"""
        if self.combined_df is None:
            print("❌ No combined data available")
            return
        
        print(f"\n📈 Summary Statistics:")
        print("=" * 50)
        print(f"Total files processed: {len(self.json_data)}")
        print(f"Total rows: {len(self.combined_df)}")
        print(f"Total columns: {len(self.combined_df.columns)}")
        
        # Show numeric columns summary
        numeric_cols = self.combined_df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            print(f"\nNumeric columns summary:")
            print(self.combined_df[numeric_cols].describe().to_string())


# Example usage functions
def example_basic_usage():
    """Example: Basic usage with multiple JSON files"""
    print("\n🔄 Example 1: Basic Multiple JSON to Single Excel")
    print("=" * 60)
    
    # Create sample JSON files for demonstration
    create_sample_json_files()
    
    converter = MultipleJSONToSingleExcelConverter()
    
    # Method 1: Load specific files
    json_files = ["employees.json", "products.json", "orders.json"]
    converter.load_multiple_json_files(json_files)
    
    # Method 2: Load from directory (alternative)
    # converter.load_from_directory(".", "*.json")
    
    # Combine data
    converter.combine_json_data(add_source_column=True)
    
    # Preview data
    converter.preview_combined_data()
    
    # Save to Excel
    converter.save_to_excel_styled("combined_data.xlsx")
    
    # Get summary stats
    converter.get_summary_stats()


def example_advanced_usage():
    """Example: Advanced usage with custom processing"""
    print("\n🔄 Example 2: Advanced Multiple JSON Processing")
    print("=" * 60)
    
    converter = MultipleJSONToSingleExcelConverter()
    
    # Load from directory
    converter.load_from_directory("./data", "*.json")
    
    # Combine without source column
    converter.combine_json_data(add_source_column=False)
    
    # Add custom processing
    if converter.combined_df is not None:
        # Add timestamp column
        converter.combined_df['processed_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add row number
        converter.combined_df['row_id'] = range(1, len(converter.combined_df) + 1)
        
        print("✅ Added custom columns: processed_date, row_id")
    
    converter.preview_combined_data()
    converter.save_to_excel_styled("advanced_combined_data.xlsx")


def create_sample_json_files():
    """Create sample JSON files for demonstration"""
    
    # Sample employees data
    employees = [
        {"id": 1, "name": "John Doe", "department": "IT", "salary": 50000, "hire_date": "2020-01-15"},
        {"id": 2, "name": "Jane Smith", "department": "HR", "salary": 55000, "hire_date": "2019-03-22"},
        {"id": 3, "name": "Bob Johnson", "department": "IT", "salary": 48000, "hire_date": "2021-06-10"}
    ]
    
    # Sample products data
    products = [
        {"product_id": "P001", "name": "Laptop", "category": "Electronics", "price": 999.99, "stock": 50},
        {"product_id": "P002", "name": "Mouse", "category": "Electronics", "price": 25.99, "stock": 200},
        {"product_id": "P003", "name": "Keyboard", "category": "Electronics", "price": 75.50, "stock": 150}
    ]
    
    # Sample orders data
    orders = [
        {"order_id": "O001", "customer": "Alice Brown", "product_id": "P001", "quantity": 1, "total": 999.99, "order_date": "2023-12-01"},
        {"order_id": "O002", "customer": "Charlie Davis", "product_id": "P002", "quantity": 2, "total": 51.98, "order_date": "2023-12-02"},
        {"order_id": "O003", "customer": "Diana Wilson", "product_id": "P003", "quantity": 1, "total": 75.50, "order_date": "2023-12-03"}
    ]
    
    # Save sample files
    with open("employees.json", "w") as f:
        json.dump(employees, f, indent=2)
    
    with open("products.json", "w") as f:
        json.dump(products, f, indent=2)
    
    with open("orders.json", "w") as f:
        json.dump(orders, f, indent=2)
    
    print("✅ Created sample JSON files: employees.json, products.json, orders.json")


# Main execution
if __name__ == "__main__":
    converter = MultipleJSONToSingleExcelConverter()
    

    converter.load_from_directory(".", "*.json")
    
    # Combine data
    converter.combine_json_data(add_source_column=True)
    
    # Preview data
    converter.preview_combined_data()
    
    # Save to Excel
    converter.save_to_excel_simple("combined_data.xlsx")
    
    # Get summary stats
    converter.get_summary_stats()